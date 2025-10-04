import psycopg2
import os
import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from pathlib import Path
from datetime import datetime
from openpyxl.styles import Font, Alignment
timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
export_path = f"exports/report_{timestamp}.xlsx"


# --- параметры подключения ---
DB_HOST = "localhost"
DB_PORT = "5432"
DB_NAME = "fashion_store"
DB_USER = "postgres"
DB_PASSWORD = "ap112005"

# --- подключение к БД ---
conn = psycopg2.connect(
    host=DB_HOST,
    port=DB_PORT,
    dbname=DB_NAME,
    user=DB_USER,
    password=DB_PASSWORD
)

# --- создаём папки charts и exports ---
Path("charts").mkdir(exist_ok=True)
Path("exports").mkdir(exist_ok=True)

# --- читаем SQL ---
with open("queries2.sql", "r", encoding="utf-8") as f:
    raw_sql = f.read()
queries = [q.strip() for q in raw_sql.split(';') if q.strip()]

# --- обработка ---

query_results = {} 

for idx, query in enumerate(queries, start=1):
    print(f"\n===== Запрос {idx} =====")
    try:
        df = pd.read_sql_query(query, conn)
        print(f"Получено строк: {len(df)}")
        print(df.head())

        if df.empty:
            print("⚠️ Результат пуст — график не будет построен.")
            continue

        query_results[f"Query_{idx}"] = df  # ← добавь это

        plt.figure(figsize=(8, 5))

        # === 1. Top-selling products (Bar chart) ===
        if idx == 1:
            df.plot.bar(x='product_name', y='total_sold', color='skyblue', legend=False)
            plt.title("Топ продаваемых товаров по количеству")
            plt.ylabel("Количество продаж")
            plt.xticks(rotation=45, ha='right')

        # === 2. Revenue by country (Pie chart) ===
        elif idx == 2:
            df.set_index('country')['total_revenue'].plot.pie(autopct='%1.1f%%')
            plt.title("Распределение выручки по странам")
            plt.ylabel("")

        # === 3. Average order value per country (Horizontal bar) ===
        elif idx == 3:
            df.plot.barh(x='country', y='avg_order_value', color='orange', legend=False)
            plt.title("Средняя стоимость заказа по странам")
            plt.xlabel("Средняя сумма заказа")

        # === 4. Average stock per category (Line chart) ===
        elif idx == 4:
            df.plot(x='category', y='avg_stock', marker='o', linestyle='-', color='green', legend=False)
            plt.title("Средний запас по категориям товаров")
            plt.ylabel("Средний уровень запаса")
            plt.xticks(rotation=45, ha='right')

        # === 5. Revenue by channel (Histogram) ===
        elif idx == 5:
            plt.bar(df['channel'], df['revenue'], color='purple')
            plt.title("Выручка по каналам продаж")
            plt.ylabel("Выручка")

        # === 6. Product profitability (Scatter) ===
        elif idx == 6:
            plt.figure(figsize=(10,6))

            plt.scatter(df['cost_price'], df['catalog_price'], s=50, alpha=0.7, color='orange')

            # Добавляем подписи товаров
            for i, row in df.iterrows():
                plt.text(row['cost_price'], row['catalog_price'], row['product_name'], fontsize=8, alpha=0.7)

            plt.title(f"Прибыльность продуктов")
            plt.xlabel("Себестоимость")
            plt.ylabel("Цена каталога")
            plt.grid(True)
            plt.tight_layout()
            plt.show()


            chart_path = f"charts/query_{idx}.png"
            plt.savefig(chart_path)
            plt.close()
            print(f"✅ Scatter-график сохранён: {chart_path}")


        elif idx == 7:
            fig = px.bar(
                df,
                x="category",
                y="total_sales",
                color="category",
                animation_frame=df["month"].astype(str),  # каждый месяц — отдельный кадр
                title="Изменение продаж по категориям со временем"
            )
            fig.update_layout(xaxis_title="Категория", yaxis_title="Общие продажи")
            fig.show()
            fig.write_html("charts/query_7.html")
            print(f"Интерактивный график успешно отображён.")
        

        
        plt.tight_layout()
        chart_path = f"charts/query_{idx}.png"
        plt.savefig(chart_path)
        plt.close()
        print(f"✅ График сохранён: {chart_path}")

    except Exception as e:
        print(f"Ошибка при выполнении запроса {idx}: {e}")


def export_to_excel(dataframes_dict, filename):
    """Экспортирует результаты всех запросов в Excel с форматированием"""
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        for sheet_name, df in dataframes_dict.items():
            # --- Убираем timezone из дат ---
            for col in df.select_dtypes(include=['datetimetz']).columns:
                df[col] = df[col].dt.tz_localize(None)

            # Добавляем индекс в датафрейм
            df_reset = df.reset_index(drop=True)
            df_reset.index += 1  # начинаем с 1
            df_reset.to_excel(writer, sheet_name=sheet_name, index=True)

    wb = load_workbook(filename)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.freeze_panes = "B2"  # закрепляем заголовки
        ws.auto_filter.ref = ws.dimensions  # включаем фильтры

        # Жирные заголовки и выравнивание
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Градиентное форматирование числовых колонок
        for col in ws.iter_cols(min_row=2):
            values = [cell.value for cell in col if isinstance(cell.value, (int, float))]
            if not values:
                continue
            col_letter = col[0].column_letter
            max_row = ws.max_row
            rule = ColorScaleRule(
                start_type="min", start_color="FFAA0000",
                mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                end_type="max", end_color="FF00AA00"
            )
            ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{max_row}", rule)

        # Авто-подгонка ширины колонок
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

    wb.save(filename)

    total_sheets = len(wb.sheetnames)
    total_rows = sum(ws.max_row - 1 for ws in wb.worksheets)  # вычитаем заголовки
    print(f"✅ Создан файл {filename}, {total_sheets} листов, {total_rows} строк.")


export_to_excel(query_results, export_path)
conn.close()
print("\n✅ Все запросы успешно выполнены и графики сохранены в папку /charts.")
